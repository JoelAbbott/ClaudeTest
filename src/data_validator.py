import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
import openpyxl
from openpyxl.styles import PatternFill
import logging

from .validation_result import ValidationResult


class DataValidator:
    """Core data validation engine for Excel and CSV files."""
    
    def __init__(self):
        """Initialize the data validator."""
        self.logger = logging.getLogger(__name__)
        
        # Color schemes for Excel output
        self.colors = {
            'error': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
            'warning': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
            'passed': PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        }
    
    def validate_file(self, file_path: str, rules: Dict[str, Any]) -> ValidationResult:
        """Validate an Excel or CSV file against specified rules.
        
        Args:
            file_path: Path to the file to validate.
            rules: Dictionary containing validation rules.
            
        Returns:
            ValidationResult object containing all validation findings.
            
        Raises:
            FileNotFoundError: If the specified file doesn't exist.
            ValueError: If the file format is not supported.
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Load the data
        try:
            if file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path)
            elif file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_path}")
        except Exception as e:
            raise ValueError(f"Error reading file {file_path}: {str(e)}")
        
        result = ValidationResult(source_file=file_path)
        
        # Check for empty dataframe
        if df.empty:
            result.add_warning('File contains no data', '', None)
            return result
        
        # Validate required columns
        if 'required_columns' in rules:
            self._validate_required_columns(df, rules['required_columns'], result)
        
        # Detect missing values
        missing_result = self.detect_missing_values(df)
        result.errors.extend(missing_result.errors)
        result.warnings.extend(missing_result.warnings)
        result.passed.extend(missing_result.passed)
        
        # Validate data types
        if 'data_types' in rules:
            type_result = self.validate_data_types(df, rules['data_types'])
            result.errors.extend(type_result.errors)
            result.warnings.extend(type_result.warnings)
            result.passed.extend(type_result.passed)
        
        return result
    
    def detect_missing_values(self, df: pd.DataFrame) -> ValidationResult:
        """Detect missing values in the dataframe.
        
        Args:
            df: Pandas DataFrame to check for missing values.
            
        Returns:
            ValidationResult containing missing value findings.
        """
        result = ValidationResult(source_file='')
        
        if df.empty:
            result.add_warning('Empty dataframe provided', '', None)
            return result
        
        for column in df.columns:
            # Check for None/NaN values
            null_mask = df[column].isnull()
            null_indices = df[null_mask].index.tolist()
            
            # Check for empty strings
            if df[column].dtype == 'object':
                empty_mask = df[column].astype(str).str.strip() == ''
                empty_indices = df[empty_mask & ~null_mask].index.tolist()
            else:
                empty_indices = []
            
            # Record findings
            for idx in null_indices:
                result.add_error(f'Missing value (null/NaN)', column, idx)
            
            for idx in empty_indices:
                result.add_error(f'Missing value (empty string)', column, idx)
            
            # Record passed validations for non-missing values
            valid_count = len(df) - len(null_indices) - len(empty_indices)
            if valid_count > 0:
                result.add_passed(f'{valid_count} valid values', column)
        
        return result
    
    def validate_data_types(self, df: pd.DataFrame, type_rules: Dict[str, str]) -> ValidationResult:
        """Validate data types against specified rules.
        
        Args:
            df: Pandas DataFrame to validate.
            type_rules: Dictionary mapping column names to expected types.
            
        Returns:
            ValidationResult containing type validation findings.
        """
        result = ValidationResult(source_file='')
        
        for column, expected_type in type_rules.items():
            if column not in df.columns:
                result.add_error(f'Column not found in data', column, None)
                continue
            
            col_data = df[column].dropna()  # Remove NaN values for type checking
            
            if col_data.empty:
                result.add_warning(f'Column contains only missing values', column, None)
                continue
            
            # Check each value against expected type
            for idx, value in col_data.items():
                is_valid = self._check_data_type(value, expected_type)
                
                if not is_valid:
                    result.add_error(
                        f'Invalid data type. Expected {expected_type}, got {type(value).__name__}',
                        column,
                        idx,
                        value
                    )
                else:
                    result.add_passed(f'Valid {expected_type} value', column, idx)
        
        return result
    
    def _validate_required_columns(self, df: pd.DataFrame, required_columns: List[str], result: ValidationResult) -> None:
        """Validate that all required columns are present.
        
        Args:
            df: Pandas DataFrame to check.
            required_columns: List of required column names.
            result: ValidationResult to update with findings.
        """
        missing_columns = set(required_columns) - set(df.columns)
        
        for column in missing_columns:
            result.add_error(f'Required column missing', column, None)
        
        for column in required_columns:
            if column in df.columns:
                result.add_passed(f'Required column present', column)
    
    def _check_data_type(self, value: Any, expected_type: str) -> bool:
        """Check if a value matches the expected data type.
        
        Args:
            value: Value to check.
            expected_type: Expected type as string ('int', 'float', 'str', etc.).
            
        Returns:
            True if value matches expected type, False otherwise.
        """
        try:
            if expected_type == 'int':
                if isinstance(value, (int, np.integer)):
                    return True
                if isinstance(value, (float, np.floating)):
                    return value.is_integer()
                if isinstance(value, str):
                    int(value)
                    return True
                return False
            
            elif expected_type == 'float':
                if isinstance(value, (int, float, np.number)):
                    return True
                if isinstance(value, str):
                    float(value)
                    return True
                return False
            
            elif expected_type == 'str':
                return isinstance(value, str)
            
            elif expected_type == 'bool':
                return isinstance(value, (bool, np.bool_))
            
            else:
                self.logger.warning(f"Unknown data type: {expected_type}")
                return True
                
        except (ValueError, TypeError):
            return False
    
    def generate_colored_output(self, result: ValidationResult, output_path: str) -> None:
        """Generate color-coded Excel output with validation results.
        
        Args:
            result: ValidationResult containing validation findings.
            output_path: Path where the output Excel file should be saved.
        """
        # Load original data
        if result.source_file.endswith('.xlsx'):
            original_df = pd.read_excel(result.source_file)
        else:
            original_df = pd.read_csv(result.source_file)
        
        # Create a copy for styling
        styled_df = original_df.copy()
        
        # Create summary data
        summary_data = []
        all_entries = result.errors + result.warnings + result.passed
        
        for entry in all_entries:
            summary_data.append({
                'Column': entry.column,
                'Row': entry.row if entry.row is not None else 'All',
                'Severity': entry.severity.title(),
                'Message': entry.message,
                'Value': entry.value if entry.value is not None else ''
            })
        
        summary_df = pd.DataFrame(summary_data)
        
        # Create lineage data
        lineage_data = [{
            'source_file': result.source_file,
            'total_rows': len(original_df),
            'total_columns': len(original_df.columns),
            'validation_timestamp': pd.Timestamp.now().isoformat(),
            'total_errors': len(result.errors),
            'total_warnings': len(result.warnings),
            'total_passed': len(result.passed)
        }]
        lineage_df = pd.DataFrame(lineage_data)
        
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write the original data (will be colored)
            styled_df.to_excel(writer, sheet_name='Validated_Data', index=False)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            lineage_df.to_excel(writer, sheet_name='Data_Lineage', index=False)
            
            # Apply color coding to the validated data sheet
            self._apply_color_coding(writer, result, 'Validated_Data')
    
    def _apply_color_coding(self, writer: pd.ExcelWriter, result: ValidationResult, sheet_name: str) -> None:
        """Apply color coding to Excel cells based on validation results.
        
        Args:
            writer: Excel writer object.
            result: ValidationResult containing validation findings.
            sheet_name: Name of the sheet to apply coloring to.
        """
        workbook = writer.book
        worksheet = workbook[sheet_name]
        
        # Create a mapping of (row, col) to severity
        cell_severity = {}
        
        for entry in result.errors + result.warnings + result.passed:
            if entry.row is not None:
                # Convert to 1-indexed for Excel (add 2 for header row)
                excel_row = entry.row + 2
                
                # Find column index
                try:
                    if sheet_name in writer.sheets:
                        df = pd.read_excel(result.source_file)
                        col_idx = list(df.columns).index(entry.column) + 1  # 1-indexed
                        cell_severity[(excel_row, col_idx)] = entry.severity
                except (ValueError, KeyError):
                    continue
        
        # Apply colors
        for (row, col), severity in cell_severity.items():
            cell = worksheet.cell(row=row, column=col)
            if severity in self.colors:
                cell.fill = self.colors[severity]